import csv
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def write_results(search_id: str, matched_jobs: list) -> tuple:
    """
    Generates both a CSV and an Excel file for the matched jobs.
    Returns (csv_filename, excel_filename).
    """
    output_dir = getattr(settings, "OUTPUT_DIR", os.path.join(settings.BASE_DIR, "output"))
    os.makedirs(output_dir, exist_ok=True)

    csv_filename = f"jobs_{search_id}.csv"
    excel_filename = f"jobs_{search_id}.xlsx"

    csv_path = os.path.join(output_dir, csv_filename)
    excel_path = os.path.join(output_dir, excel_filename)

    # 1. Write CSV
    headers = ["Rank", "Job Title", "Company", "Location", "Salary", "Match Score", "Why It's a Match", "Source", "Apply URL"]
    with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for idx, job in enumerate(matched_jobs):
            writer.writerow([
                idx + 1,
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                job.get("salary", ""),
                job.get("score", 7),
                job.get("reason", ""),
                job.get("source", ""),
                job.get("url", "")
            ])

    # 2. Write Excel (.xlsx) with professional styles
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Jobs"

    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True

    # Write Headers
    ws.append(headers)

    # Style Header Row
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark steel blue
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Add Data
    for idx, job in enumerate(matched_jobs):
        rank = idx + 1
        row_data = [
            rank,
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("salary", ""),
            job.get("score", 7),
            job.get("reason", ""),
            job.get("source", ""),
            job.get("url", "")
        ]
        ws.append(row_data)

    # Style Data Cells
    font_regular = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Row styles and formulas
    for row_idx in range(2, len(matched_jobs) + 3):
        # Rank column (center)
        ws.cell(row=row_idx, column=1).font = font_bold
        ws.cell(row=row_idx, column=1).alignment = align_center
        
        # Text columns
        for col_idx in (2, 3, 4, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_regular
            cell.alignment = align_left

        # Match Score (center, bold)
        ws.cell(row=row_idx, column=6).font = font_bold
        ws.cell(row=row_idx, column=6).alignment = align_center

        # Match Reason
        ws.cell(row=row_idx, column=7).font = font_regular
        ws.cell(row=row_idx, column=7).alignment = align_left

        # Source
        ws.cell(row=row_idx, column=8).font = font_regular
        ws.cell(row=row_idx, column=8).alignment = align_center

        # Apply URL with hyperlink formula
        url_cell = ws.cell(row=row_idx, column=9)
        url = url_cell.value
        if url:
            url_cell.value = f'=HYPERLINK("{url}", "Apply Now")'
            url_cell.font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
            url_cell.alignment = align_center
        else:
            url_cell.font = font_regular
            url_cell.alignment = align_center

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            val = str(cell.value or '')
            if val.startswith('='):  # Ignore formula length
                val = "Apply Now"
            max_len = max(max_len, len(val))
            
        # Give some padding
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

    # Set row heights
    ws.row_dimensions[1].height = 28
    for row_idx in range(2, len(matched_jobs) + 3):
        ws.row_dimensions[row_idx].height = 22

    wb.save(excel_path)
    return csv_filename, excel_filename
